from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, Dict, Any
from datetime import datetime
from io import BytesIO
import json

from app.database import get_db
from app.models import (
    Tournament, TournamentPlayer, Player, Match, MatchPlayer,
    Room, Foul, ScoreChangeLog, Notification, Substitution,
    Group, TournamentReferee, MatchStatus, ResultType, TournamentStatus
)
from app.schemas import ExportFinalResultResponse, CleanTestDataResponse
from app.utils.scoring import calculate_tournament_rankings

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

router = APIRouter(prefix="/admin", tags=["管理与导出"])


@router.get("/tournaments/{tournament_id}/export/final", response_model=ExportFinalResultResponse)
def export_final_results(
    tournament_id: int,
    db: Session = Depends(get_db)
):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(status_code=404, detail="赛事不存在")

    if tournament.status not in [TournamentStatus.FINISHED.value, TournamentStatus.ONGOING.value]:
        raise HTTPException(status_code=400, detail="赛事未结束，无法导出最终成绩")

    rankings = calculate_tournament_rankings(db, tournament_id)

    reg_map = {}
    for r in rankings:
        reg = db.query(TournamentPlayer).filter(
            TournamentPlayer.id == r["tournament_player_id"]
        ).first()
        if reg:
            reg_map[reg.id] = reg
            if not reg.final_rank:
                reg.final_rank = r["rank"]
    db.commit()

    all_registrations = db.query(TournamentPlayer).filter(
        TournamentPlayer.tournament_id == tournament_id
    ).all()

    player_details = []
    for r in rankings:
        reg = reg_map.get(r["tournament_player_id"])
        player = db.query(Player).filter(Player.id == r["player_id"]).first()

        player_matches = db.query(MatchPlayer).filter(
            MatchPlayer.tournament_player_id == r["tournament_player_id"]
        ).all()

        match_records = []
        for pm in player_matches:
            m = db.query(Match).filter(Match.id == pm.match_id).first()
            opponents = db.query(MatchPlayer).filter(
                MatchPlayer.match_id == pm.match_id,
                MatchPlayer.player_id != r["player_id"]
            ).all()
            match_records.append({
                "round": m.round_number if m else None,
                "score": pm.score,
                "result": pm.result,
                "tiebreaker": pm.tiebreaker_score,
                "opponents": [
                    {
                        "name": opp.player.name if opp.player else "",
                        "score": opp.score,
                        "result": opp.result,
                    }
                    for opp in opponents
                ],
            })

        player_fouls = db.query(Foul).filter(
            Foul.tournament_id == tournament_id,
            Foul.player_id == r["player_id"],
            Foul.approved == True
        ).all()

        player_details.append({
            "rank": r["rank"],
            "final_rank": reg.final_rank if reg else r["rank"],
            "player_id": r["player_id"],
            "registration_id": r["tournament_player_id"],
            "player_name": r["player_name"],
            "team": r.get("team"),
            "phone": player.phone if player else None,
            "email": player.email if player else None,
            "id_card": player.id_card if player else None,
            "played_rounds": r["played_rounds"],
            "wins": r["wins"],
            "draws": r["draws"],
            "losses": r["losses"],
            "win_rate": round(r["wins"] / r["played_rounds"] * 100, 2) if r["played_rounds"] > 0 else 0,
            "match_points": r["match_points"],
            "opponent_match_win_percent": r["opponent_match_win_percent"],
            "game_win_percent": r["game_win_percent"],
            "opponent_game_win_percent": r["opponent_game_win_percent"],
            "total_tiebreaker": r["total_tiebreaker"],
            "foul_penalty": r["foul_penalty"],
            "foul_count": len(player_fouls),
            "foul_details": [
                {
                    "type": f.type,
                    "description": f.description,
                    "penalty_points": f.penalty_points,
                    "match_id": f.match_id,
                }
                for f in player_fouls
            ],
            "is_substitute": reg.is_substitute if reg else False,
            "drop_round": reg.drop_round if reg else None,
            "checked_in": reg.checked_in if reg else False,
            "match_records": match_records,
        })

    match_summary = []
    all_matches = db.query(Match).filter(Match.tournament_id == tournament_id).all()
    for m in all_matches:
        room = db.query(Room).filter(Room.id == m.room_id).first()
        mps = db.query(MatchPlayer).filter(MatchPlayer.match_id == m.id).all()
        match_summary.append({
            "match_id": m.id,
            "round_number": m.round_number,
            "room_id": m.room_id,
            "room_number": room.room_number if room else m.table_number,
            "table_number": m.table_number,
            "status": m.status,
            "start_time": m.start_time,
            "end_time": m.end_time,
            "submitted_by": m.submitted_by,
            "referee_note": m.referee_note,
            "players": [
                {
                    "name": mp.player.name if mp.player else "",
                    "seat": mp.seat_position,
                    "score": mp.score,
                    "result": mp.result,
                    "is_winner": mp.is_winner,
                }
                for mp in mps
            ],
        })

    substitutions = db.query(Substitution).filter(
        Substitution.tournament_id == tournament_id
    ).all()

    change_logs = db.query(ScoreChangeLog).join(Match).filter(
        Match.tournament_id == tournament_id
    ).all()

    data = {
        "tournament": {
            "id": tournament.id,
            "name": tournament.name,
            "description": tournament.description,
            "game_type": tournament.game_type,
            "start_date": tournament.start_date,
            "end_date": tournament.end_date,
            "location": tournament.location,
            "organizer": tournament.organizer,
            "total_rounds": tournament.total_rounds,
            "max_players": tournament.max_players,
            "players_per_room": tournament.players_per_room,
            "win_points": tournament.win_points,
            "draw_points": tournament.draw_points,
            "lose_points": tournament.lose_points,
            "status": tournament.status,
        },
        "statistics": {
            "total_registered": len(all_registrations),
            "total_checked_in": sum(1 for r in all_registrations if r.checked_in),
            "total_matches": len(all_matches),
            "completed_matches": sum(
                1 for m in all_matches
                if m.status in [MatchStatus.FINISHED.value, MatchStatus.OVERRIDDEN.value]
            ),
            "total_substitutions": len(substitutions),
            "total_score_changes": len(change_logs),
        },
        "final_rankings": player_details,
        "all_matches": match_summary,
        "substitutions": [
            {
                "id": s.id,
                "old_player_id": s.old_player_id,
                "old_player_name": s.old_player.name if s.old_player else "",
                "new_player_id": s.new_player_id,
                "new_player_name": s.new_player.name if s.new_player else "",
                "effective_round": s.effective_round,
                "reason": s.reason,
                "approved_by": s.approved_by,
                "created_at": s.created_at,
            }
            for s in substitutions
        ],
        "referee_audit": [
            {
                "log_id": cl.id,
                "match_id": cl.match_id,
                "referee_id": cl.referee_id,
                "change_type": cl.change_type,
                "reason": cl.reason,
                "created_at": cl.created_at,
            }
            for cl in change_logs
        ],
    }

    return ExportFinalResultResponse(
        tournament_id=tournament_id,
        tournament_name=tournament.name,
        export_time=datetime.now(),
        data=data
    )


@router.get("/tournaments/{tournament_id}/export/excel")
def export_excel(
    tournament_id: int,
    db: Session = Depends(get_db)
):
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel导出功能不可用，请安装openpyxl库")

    result = export_final_results(tournament_id, db)
    data = result.data

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "最终排名"

    headers = ["排名", "选手姓名", "所属队伍", "参赛轮次", "胜", "平", "负", "胜率%",
               "积分", "对手胜率%", "对局胜率%", "对手对局胜率%", "小分", "犯规扣分", "是否替补", "退赛轮次"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    rankings = data["final_rankings"]
    for row_idx, player in enumerate(rankings, 2):
        values = [
            player["rank"],
            player["player_name"],
            player["team"] or "",
            player["played_rounds"],
            player["wins"],
            player["draws"],
            player["losses"],
            player["win_rate"],
            player["match_points"],
            round(player["opponent_match_win_percent"] * 100, 2),
            round(player["game_win_percent"] * 100, 2),
            round(player["opponent_game_win_percent"] * 100, 2),
            player["total_tiebreaker"],
            player["foul_penalty"],
            "是" if player["is_substitute"] else "否",
            player["drop_round"] or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        if row_idx <= 4:
            medal_fill = PatternFill(
                start_color=["FFD700", "C0C0C0", "CD7F32"][row_idx - 2] if row_idx - 2 < 3 else "E2EFDA",
                end_color=["FFD700", "C0C0C0", "CD7F32"][row_idx - 2] if row_idx - 2 < 3 else "E2EFDA",
                fill_type="solid"
            )
            for col in range(1, len(headers) + 1):
                ws1.cell(row=row_idx, column=col).fill = medal_fill

    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[chr(64 + col)].width = 15

    ws2 = wb.create_sheet("赛事信息")
    t = data["tournament"]
    info = [
        ["赛事名称", t["name"]],
        ["赛事类型", t["game_type"]],
        ["主办方", t["organizer"] or ""],
        ["开始时间", str(t["start_date"]) if t["start_date"] else ""],
        ["结束时间", str(t["end_date"]) if t["end_date"] else ""],
        ["比赛地点", t["location"] or ""],
        ["总轮次", t["total_rounds"]],
        ["最大人数", t["max_players"]],
        ["每桌人数", t["players_per_room"]],
        ["胜场积分", t["win_points"]],
        ["平局积分", t["draw_points"]],
        ["负场积分", t["lose_points"]],
        ["赛事状态", t["status"]],
        ["赛事描述", t["description"] or ""],
    ]
    for row, (key, val) in enumerate(info, 1):
        ws2.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=val)
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 50

    stats = data["statistics"]
    ws2.cell(row=17, column=1, value="统计信息").font = Font(bold=True, size=12)
    stat_row = 18
    for k, v in stats.items():
        ws2.cell(row=stat_row, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=stat_row, column=2, value=v)
        stat_row += 1

    ws3 = wb.create_sheet("对局记录")
    match_headers = ["对局ID", "轮次", "桌号", "状态", "开始时间", "结束时间", "提交人",
                     "选手1", "分数", "结果", "选手2", "分数", "结果",
                     "选手3", "分数", "结果", "选手4", "分数", "结果"]
    for col, h in enumerate(match_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, match in enumerate(data["all_matches"], 2):
        values = [
            match["match_id"],
            match["round_number"],
            match["room_number"],
            match["status"],
            str(match["start_time"]) if match["start_time"] else "",
            str(match["end_time"]) if match["end_time"] else "",
            match["submitted_by"] or "",
        ]
        for i in range(4):
            if i < len(match["players"]):
                p = match["players"][i]
                values.extend([p["name"], p["score"], p["result"]])
            else:
                values.extend(["", "", ""])
        for col, val in enumerate(values, 1):
            ws3.cell(row=row_idx, column=col, value=val)

    for col in range(1, len(match_headers) + 1):
        ws3.column_dimensions[chr(64 + col)].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{result.tournament_name}_最终成绩_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.post("/clean-test-data", response_model=CleanTestDataResponse)
def clean_test_data(
    tournament_id: Optional[int] = None,
    clean_all_test: bool = False,
    db: Session = Depends(get_db)
):
    deleted_counts = {
        "tournaments": 0,
        "groups": 0,
        "rooms": 0,
        "tournament_players": 0,
        "matches": 0,
        "match_players": 0,
        "fouls": 0,
        "notifications": 0,
        "substitutions": 0,
        "tournament_referees": 0,
        "score_change_logs": 0,
    }

    if tournament_id:
        tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
        if not tournament:
            raise HTTPException(status_code=404, detail="赛事不存在")
        tournaments_to_delete = [tournament]
    elif clean_all_test:
        tournaments_to_delete = db.query(Tournament).filter(
            Tournament.is_test == True
        ).all()
    else:
        raise HTTPException(
            status_code=400,
            detail="请指定 tournament_id 或设置 clean_all_test=true"
        )

    tournament_ids = [t.id for t in tournaments_to_delete]

    for tid in tournament_ids:
        match_ids = [m.id for m in db.query(Match.id).filter(Match.tournament_id == tid).all()]

        if match_ids:
            deleted_counts["score_change_logs"] += db.query(ScoreChangeLog).filter(
                ScoreChangeLog.match_id.in_(match_ids)
            ).delete(synchronize_session=False)
            deleted_counts["fouls"] += db.query(Foul).filter(
                Foul.match_id.in_(match_ids)
            ).delete(synchronize_session=False)
            deleted_counts["match_players"] += db.query(MatchPlayer).filter(
                MatchPlayer.match_id.in_(match_ids)
            ).delete(synchronize_session=False)
            deleted_counts["matches"] += db.query(Match).filter(
                Match.id.in_(match_ids)
            ).delete(synchronize_session=False)

        deleted_counts["fouls"] += db.query(Foul).filter(
            Foul.tournament_id == tid, Foul.match_id.is_(None)
        ).delete(synchronize_session=False)
        deleted_counts["notifications"] += db.query(Notification).filter(
            Notification.tournament_id == tid
        ).delete(synchronize_session=False)
        deleted_counts["substitutions"] += db.query(Substitution).filter(
            Substitution.tournament_id == tid
        ).delete(synchronize_session=False)
        deleted_counts["tournament_referees"] += db.query(TournamentReferee).filter(
            TournamentReferee.tournament_id == tid
        ).delete(synchronize_session=False)
        deleted_counts["tournament_players"] += db.query(TournamentPlayer).filter(
            TournamentPlayer.tournament_id == tid
        ).delete(synchronize_session=False)
        deleted_counts["rooms"] += db.query(Room).filter(
            Room.tournament_id == tid
        ).delete(synchronize_session=False)
        deleted_counts["groups"] += db.query(Group).filter(
            Group.tournament_id == tid
        ).delete(synchronize_session=False)

    deleted_counts["tournaments"] += db.query(Tournament).filter(
        Tournament.id.in_(tournament_ids)
    ).delete(synchronize_session=False)

    db.commit()

    total_deleted = sum(deleted_counts.values())

    return CleanTestDataResponse(
        tournament_id=tournament_id,
        deleted_counts=deleted_counts,
        message=f"测试数据清理完成，共删除 {total_deleted} 条记录，涉及 {len(tournament_ids)} 个赛事"
    )


@router.get("/tournaments/{tournament_id}/summary")
def get_tournament_admin_summary(
    tournament_id: int,
    db: Session = Depends(get_db)
):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(status_code=404, detail="赛事不存在")

    all_regs = db.query(TournamentPlayer).filter(
        TournamentPlayer.tournament_id == tournament_id
    ).all()
    subs = [r for r in all_regs if r.is_substitute]
    regular = [r for r in all_regs if not r.is_substitute]
    checked_in = [r for r in all_regs if r.checked_in]
    dropped = [r for r in all_regs if r.drop_round is not None]

    all_matches = db.query(Match).filter(Match.tournament_id == tournament_id).all()
    rounds_data = {}
    for m in all_matches:
        if m.round_number not in rounds_data:
            rounds_data[m.round_number] = {
                "round": m.round_number,
                "total": 0,
                "pending": 0,
                "locked": 0,
                "ongoing": 0,
                "finished": 0,
                "cancelled": 0,
                "overridden": 0,
            }
        rounds_data[m.round_number]["total"] += 1
        key = m.status
        if key in rounds_data[m.round_number]:
            rounds_data[m.round_number][key] += 1

    groups = db.query(Group).filter(Group.tournament_id == tournament_id).all()
    rooms = db.query(Room).filter(Room.tournament_id == tournament_id).all()
    refs = db.query(TournamentReferee).filter(
        TournamentReferee.tournament_id == tournament_id
    ).all()
    fouls = db.query(Foul).filter(
        Foul.tournament_id == tournament_id
    ).all()
    substitutions = db.query(Substitution).filter(
        Substitution.tournament_id == tournament_id
    ).all()
    changes = db.query(ScoreChangeLog).join(Match).filter(
        Match.tournament_id == tournament_id
    ).all()

    return {
        "tournament": {
            "id": tournament.id,
            "name": tournament.name,
            "game_type": tournament.game_type,
            "status": tournament.status,
            "is_test": tournament.is_test,
            "total_rounds": tournament.total_rounds,
        },
        "registration": {
            "total": len(all_regs),
            "regular_players": len(regular),
            "substitutes": len(subs),
            "checked_in": len(checked_in),
            "checked_in_rate": round(len(checked_in) / len(all_regs) * 100, 2) if all_regs else 0,
            "dropped_out": len(dropped),
            "max_players": tournament.max_players,
            "fill_rate": round(len(regular) / tournament.max_players * 100, 2) if tournament.max_players else 0,
        },
        "matches": {
            "total": len(all_matches),
            "rounds": sorted(rounds_data.values(), key=lambda x: x["round"]),
            "by_status": {
                "pending": sum(r["pending"] for r in rounds_data.values()),
                "locked": sum(r["locked"] for r in rounds_data.values()),
                "ongoing": sum(r["ongoing"] for r in rounds_data.values()),
                "finished": sum(r["finished"] for r in rounds_data.values()),
                "cancelled": sum(r["cancelled"] for r in rounds_data.values()),
                "overridden": sum(r["overridden"] for r in rounds_data.values()),
            },
            "completion_rate": round(
                (sum(r["finished"] + r["overridden"] for r in rounds_data.values())) / len(all_matches) * 100, 2
            ) if all_matches else 0,
        },
        "infrastructure": {
            "groups_count": len(groups),
            "rooms_count": len(rooms),
            "referees_count": len(refs),
        },
        "exceptions": {
            "fouls_count": len(fouls),
            "approved_fouls": sum(1 for f in fouls if f.approved),
            "substitutions_count": len(substitutions),
            "score_changes_count": len(changes),
        },
    }
